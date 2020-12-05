# !/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2020/12/4 下午2:34
# @Author : tangtain
# @File : order_many.py
# @Software: PyCharm


import requests
from config.config_path import *
from openpyxl import load_workbook
from scripts.get_excel_order_data import get_excel_data
from scripts.get_token import get_header_token
import json


def order_many():
    method = 'post'
    url = dev_order_url
    header = get_header_token()

    sh = get_excel_data(excel_data_path, 'order_many')
    for i in range(1, sh.max_row + 1):
        data = json.loads(sh.cell(i, 1).value)
        res = requests.request(method=method, url=url, data=data, headers=header)
        print(res.status_code)


if __name__ == '__main__':
    order_many()
